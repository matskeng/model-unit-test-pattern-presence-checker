import argparse
from pathlib import Path
from openpyxl import load_workbook

TARGET_SHEET_NAME = "テストパターン"
TARGET_CELL = "A5"
TARGET_FILENAME_PREFIX = "単体テストファイル_"


def has_test_pattern(xlsx_path: Path) -> str:
    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception:
        return "読込失敗"

    if TARGET_SHEET_NAME not in wb.sheetnames:
        return "シートなし"

    sheet = wb[TARGET_SHEET_NAME]
    value = sheet[TARGET_CELL].value

    return "あり" if value not in (None, "") else "なし"


def main(input_dir: Path, output_file: Path):
    rows = []

    for path in input_dir.rglob("*.xlsm"):
        if not path.name.startswith(TARGET_FILENAME_PREFIX):
            continue

        result = has_test_pattern(path)
        rows.append([str(path), path.name, result])

    with output_file.open("w", encoding="utf-8-sig", newline="") as f:
        f.write("ファイルパス,ファイル名,単体テストパターン有無\n")
        for row in rows:
            f.write(",".join(row) + "\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="モデル単体テストパターンの有無を確認するツール"
    )
    parser.add_argument("input_dir", help="走査対象のフォルダパス")
    parser.add_argument(
        "-o",
        "--output",
        default="pattern_check.csv",
        help="出力CSVファイル名（省略時: pattern_check.csv）"
    )

    args = parser.parse_args()
    main(Path(args.input_dir), Path(args.output))